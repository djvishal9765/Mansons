# -----Import
from flask import Flask, render_template, request, jsonify, send_file, abort
from array import array
import csv
from fileinput import close, filename
from pickletools import decimalnl_long
from unicodedata import decimal
from unittest import result
import time
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
import functools
from datetime import date
from yyyymmdd import Date
import glob
import os
import threading
from colorama import init, Fore, Back, Style
from csv import reader
import csv
import json
import re
import socket
import logging
from openpyxl import load_workbook
from io import BytesIO
import base64
from openpyxl.drawing.image import Image as OpenpyxlImage


# Initialize colorama
init(autoreset=True)
# Flask App Initialization
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'UI/'  # Folder where files are stored
# global variable
STOPThread = False
jsonFilePath = r"src/config.json"
pages =[]

def setup_logging():
    """Sets up logging configuration."""
    log_dir = "LOG"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    log_filename = os.path.join(log_dir, f"log_{datetime.now().strftime('%Y-%m-%d')}.txt")
    logging.basicConfig(
        filename=log_filename,
        level=logging.DEBUG,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    return log_filename

def read_config(file_path):
    """Reads the configuration file and returns the lines."""
    if not os.path.exists(file_path):
        logging.error(f"Config file not found: {file_path}")
        raise FileNotFoundError(f"Config file not found: {file_path}")
    with open(file_path, "r") as file:
        return [line.strip() for line in file if line.strip() and not line.startswith("#")]

def get_value(key, lines, default=None):
    """Extracts the value for a given key in the config lines."""
    for line in lines:
        if key in line:
            # Handle keys in the form `key=value` or `key: value`
            if "=" in line:
                value = line.split("=", 1)[-1].strip()
            elif ":" in line:
                value = line.split(":", 1)[-1].strip()
            else:
                continue

            if value == "":
                raise ValueError(f"Key '{key}' is missing a value in the config file.")
            return value
    if default is not None:
        return default
    raise ValueError(f"Key '{key}' not found in the config file.")

def parse_counters(lines, no_of_counter):
    """Parses counters using the specific format in the config file."""
    counters = []
    for i in range(1, no_of_counter + 1):
        cindex_line = f"CIndex={i}"
        cid_line = f"CID={i}"
        
        cindex = next((line.split("=")[1].strip() for line in lines if line.startswith(cindex_line)), None)
        cid = next((line.split("=")[1].strip() for line in lines if line.startswith(cid_line)), None)
        
        if cindex is None or cid is None:
            raise ValueError(f"Missing counter settings for CIndex={i} or CID={i}.")
        
        counters.append((int(cindex), int(cid)))
    return counters

def parse_config(lines):
    """Parses the config lines and organizes data into variables."""
    global pages
    try:
        # Version and Company Info
        version = get_value("Version", lines, "Unknown")
        company_name = get_value("Company_name", lines, "Unknown")

        # Counter Settings
        no_of_counter = int(get_value("No_of_counter", lines))
        counters = parse_counters(lines, no_of_counter)

        # CSV Settings
        no_of_csv = int(get_value("No of Files to be Crunched", lines))
        csv_paths = [get_value(f"M{i}", lines) for i in range(1, no_of_csv + 1)]

        # Display Settings
        no_of_lines = int(get_value("No. of Lines", lines))
        lines_data = [get_value(f"L{i}", lines) for i in range(1, no_of_lines + 1)]

        log_interval = int(get_value("Log Inteval (in mins)", lines))

        # Page Control
        page_enable = int(get_value("PAGE Enable (1-enable , 0-Disable)", lines))
        # page_interval = int(get_value("Page Interval (in sec)", lines))
        no_of_pages = int(get_value("No of Pages", lines))
        pages = [get_value(f"P{i}", lines) for i in range(1, no_of_pages + 1)]

        # Return parsed data
        return {
            "version": version,
            "company_name": company_name,
            "counters": counters,
            "csv_paths": csv_paths,
            "no_of_lines": no_of_lines,
            "lines_data": lines_data,
            "log_interval": log_interval,
            "page_enable": page_enable,
            # "page_interval": page_interval,
            "pages": pages,
        }
    except ValueError as e:
        logging.error(f"Configuration Error: {e}")
        raise
    except Exception as e:
        logging.error(f"Unexpected Error: {e}")
        raise


def convert_float(decimal_places, value):
    format_string = "{:." + str(decimal_places) + "f}"
    result = format_string.format(value)
    return float(result)


# ---CSV crunching
def CSV_chrunching(value, csv_paths):  # Accept csv_paths as an argument
    csv_temp = value
    csv_list = [None] * 6
    cnt = 0
    StartBit = 0
    temp = []
    for b in re.finditer(",", csv_temp):
        csv_list[cnt] = csv_temp[StartBit : b.start()]
        StartBit = b.start() + 1
        cnt = cnt + 1
    cnt = 0
    # ---File Path
    path_no = int(csv_list[0])
    csv_path = csv_paths[path_no - 1]  # Use csv_paths passed into the function
    scan = int(csv_list[1])
    cell_capture_row = int(csv_list[2])
    cell_capture_col = int(csv_list[3])
    Type = str(csv_list[4])

    # Check if the path includes a specific file name or is a folder
    if os.path.isfile(csv_path):
        # Specific file provided
        pathvalue = csv_path
        # print(Fore.GREEN + f"Specific file provided: {csv_path}" + Style.RESET_ALL)
    else:
        # Folder path provided; find the most recently modified CSV file
        file_type = "\*csv"
        files = glob.glob(csv_path + file_type)
        if not files:
            print(Fore.RED + f"No CSV files found in folder: {csv_path}" + Style.RESET_ALL)
            return -786  # Return error code for no files

        # Select the most recently modified file
        pathvalue = max(files, key=os.path.getctime)
        # print(Fore.GREEN + f"Using the most recently modified file: {pathvalue}" + Style.RESET_ALL)


    # ---Read entire file
    with open(pathvalue, encoding="utf-8") as csvf:
        csvreader = csv.reader(csvf)
        for row in csvreader:
            temp.append(row)

        if scan == 1:
            scan_row = cell_capture_row
        elif scan == 2:
            scan_row = len(temp) - 1
        elif scan == 3:
            scan_row = cell_capture_row
            scan_col = cell_capture_col

        if len(temp) == 0:
            print(Fore.BLUE + "===========FILE IS EMPTY=============")
            return -786
        else:
            Row_data = temp[scan_row]
            # print("Row Data:")
            # print(*Row_data, sep=", ")

            if len(Row_data) > cell_capture_col:
                Cell_data = Row_data[cell_capture_col]
                # print("Cell Data: " + str(Cell_data))
            else:
                print(
                    Fore.BLUE
                    + f"Error: The list Row_data does not have an element at index {cell_capture_col}."
                )
                return -786

            no_of_col = len(Row_data)

            if Type == "S":
                data_update = str(Cell_data)
            elif Type == "I":
                data_update = str(int(Cell_data))
            else:
                converted_value = convert_float(Type, float(Cell_data))
                data_update = converted_value
            return data_update

def Independent(config_data):
    global STOPThread
    global my_thread
    if STOPThread:
        print(Fore.RED + "******************************")
        print("******SOFTWARE TERMINATED*****")
        print("******************************")
        print("REASON: Due to keyBoard interrupt occurred")
        print(Style.RESET_ALL)
    else:
        cnt = 0
        cnt2 = 0
        temp_Lines = ""
        temp_var = ""
        Chrunching = ""
        max_crunch = 50
        blank_ret_check = -786
        jsonDic = {}
        List_of_crunch = [None] * max_crunch
        Replace_crunch = [None] * max_crunch
        
        # Read the existing JSON data to retain non-zero counter values
        existing_data = {}
        try:
            with open(jsonFilePath, "r", encoding="utf-8") as jsonf:
                existing_data = json.load(jsonf).get("result", {})
        except FileNotFoundError:
            print(Fore.YELLOW + "JSON file not found. A new one will be created." + Style.RESET_ALL)
        except Exception as e:
            print(Fore.RED + f"Error reading JSON file: {str(e)}" + Style.RESET_ALL)
            

        for i in range(0, config_data["no_of_lines"]):  # Use config_data["no_of_lines"]
            temp_Lines = config_data["lines_data"][i]
            indexSt = temp_Lines.find("{{")
            if indexSt != -1:
                for m in re.finditer("{{", temp_Lines):
                    temp_Lines2 = temp_Lines[m.start() : len(temp_Lines)]
                    for n in re.finditer("}}", temp_Lines2):
                        Crunching = temp_Lines2[0 : n.start() + 2]
                        List_of_crunch[cnt] = Crunching
                        cnt = cnt + 1
                        break

                no_of_crunch = cnt
                cnt = 0

                for a in range(0, no_of_crunch):
                    temp_crunching = List_of_crunch[a]
                    temp_data = temp_crunching[2 : len(temp_crunching) - 2]

                    if (
                        temp_data == "DD/MM/YY"
                        or temp_data == "DD/MM/YYYY"
                        or temp_data == "MM/DD/YY"
                        or temp_data == "MM/DD/YYYY"
                    ):
                        today = date.today()
                        curr_year = str(today.year)
                        curr_mon = str(today.month)
                        if len(curr_mon) == 1:
                            curr_mon = "0" + curr_mon
                        curr_date = today.day

                        if temp_data == "DD/MM/YY":
                            format_date = (
                                str(curr_date)
                                + "/"
                                + curr_mon
                                + "/"
                                + str(curr_year[2:4])
                            )
                            Replace_crunch[a] = format_date

                        if temp_data == "MM/DD/YY":
                            format_date = (
                                curr_mon
                                + "/"
                                + str(curr_date)
                                + "/"
                                + str(curr_year[2:4])
                            )
                            Replace_crunch[a] = format_date

                        if temp_data == "DD/MM/YYYY":
                            format_date = (
                                str(curr_date) + "/" + curr_mon + "/" + str(curr_year)
                            )
                            Replace_crunch[a] = format_date

                        if temp_data == "MM/DD/YYYY":
                            format_date = (
                                curr_mon + "/" + str(curr_date) + "/" + str(curr_year)
                            )
                            Replace_crunch[a] = format_date

                    elif temp_data == "HH:MM" or temp_data == "HH:MM:SS":
                        myobj = datetime.now()
                        curr_hr = myobj.hour
                        curr_min = myobj.minute
                        curr_sec = myobj.second
                        if temp_data == "HH:MM":
                            format_time = str(curr_hr) + ":" + str(curr_min)
                            Replace_crunch[a] = format_time
                        if temp_data == "HH:MM:SS":
                            format_time = (
                                str(curr_hr) + ":" + str(curr_min) + ":" + str(curr_sec)
                            )
                            Replace_crunch[a] = format_time
                    elif temp_data == "BLANK":
                        format_time = "  "
                        Replace_crunch[a] = format_time
                    else:
                        demo = CSV_chrunching(temp_data, config_data["csv_paths"])  # Pass csv_paths
                        if demo != blank_ret_check:
                            Replace_crunch[a] = demo
                        else:
                            print(
                                Fore.BLUE
                                + "=====SKIP: BLANK FILE FOUND, WRITING 000 AS DEFAULT====="
                            )
                            Replace_crunch[a] = "000"
                Final = temp_Lines
                for b in range(0, no_of_crunch):
                    Final = Final.replace(str(List_of_crunch[b]), str(Replace_crunch[b]))

                Replace_crunch.clear()
                List_of_crunch.clear()
                List_of_crunch = [None] * max_crunch
                Replace_crunch = [None] * max_crunch
                jsonDic["line" + str(i + 1)] = Final

            else:
                jsonDic["line" + str(i + 1)] = temp_Lines

# Add or update counters in the JSON output
        for idx, (cindex, cid) in enumerate(config_data["counters"], 1):
            if cid > 0:
                current_value = existing_data.get(str(cid), 0)
                jsonDic[str(cid)] = current_value  # Keep existing value if it's non-zero
                    
        with open(jsonFilePath, "w", encoding="utf-8") as jsonf:
            jsonString = json.dumps(dict({"result": jsonDic}), indent=4)
            jsonf.write(jsonString)

        my_thread = threading.Timer(1, functools.partial(Independent, config_data=config_data)).start()


# Function to load config data
def load_config():
    config_file_path = os.path.join('src', 'config.json')
    try:
        with open(config_file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Config file not found at {config_file_path}")
        return {}
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON from {config_file_path}: {e}")
        return {}

@app.route('/check-file', methods=['GET'])
def check_file():
    # File name (default example, dynamically change as needed)
    file_name = "UI.xlsx"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)

    if not os.path.exists(file_path):
        return f"File '{file_name}' does not exist on the server.", 404

    # Pagination duration in seconds (you can dynamically pass it via a query parameter)
    pagination_duration = request.args.get('duration', 5, type=int)

    try:
        # Load Excel using openpyxl to handle merged cells
        workbook = load_workbook(file_path, data_only=True)
        sheet_data = []

        # Load config data from the config.json file
        config_data = load_config()

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Start building HTML for the table with styling
            table_html = '<table class="table table-bordered" style="border-collapse: collapse;">'

            # Identify rows and columns that contain data
            non_empty_rows = []
            non_empty_columns = set()  # Using a set to avoid duplicate entries

            # Iterate over rows and check if they contain any data
            for row in sheet.iter_rows():
                if any(cell.value not in [None, ''] for cell in row):  # Check if the row contains any data
                    non_empty_rows.append(row[0].row)  # Store row index if it has data

            # Identify non-empty columns
            for col in sheet.iter_cols():
                if any(cell.value not in [None, ''] for cell in col):  # Check if the column contains any data
                    col_index = col[0].column
                    non_empty_columns.add(col_index)  # Add the column index

            # Generate HTML for each row and column based on non-empty ones
            for row in sheet.iter_rows(min_row=min(non_empty_rows), max_row=max(non_empty_rows)):
                row_html = '<tr>'
                for cell in row:
                    if cell.column in non_empty_columns:  # Only include non-empty columns
                        # Extract cell value
                        value = cell.value

                        # Check if the value is in the format {{key}} and replace it
                        if isinstance(value, str) and value.startswith("{{") and value.endswith("}}"):
                            temp_val = str(value)
                            key = value[2:-2].strip()  # Extract the key inside {{}} 
                            temp_key = str(key)
                            replacement_value = config_data.get("result", {}).get(temp_key, value)  # Get from config.json or keep original
                            value = str(replacement_value)
                            
                            # Check for mathematical expressions like {{1}}{{+}}{{2}}
                            # Handle concatenation and operations
                            value = handle_math_operations(value, config_data)
                            value = evaluate_expression(value)

                        # Replace None with a blank space
                        cell_value = value if value is not None else " "

                        # Extract cell style
                        font = cell.font
                        fill = cell.fill
                        alignment = cell.alignment

                        # Handle all fill color types (rgb, indexed, theme, etc.)
                        if fill.fgColor:
                            if fill.fgColor.type == "rgb":  # RGB color type
                                bg_color = f"#{fill.fgColor.rgb[2:]}"  # Extract RGB value (excluding first two "FF")
                            elif fill.fgColor.type == "indexed":  # Indexed color type
                                bg_color = fill.fgColor.indexed
                            elif fill.fgColor.type == "theme":  # Theme color type
                                theme_color = fill.fgColor.theme
                                bg_color = f"theme-{theme_color}"  # Applying default "theme" label
                            else:
                                bg_color = "#FFFFFF"
                        else:
                            bg_color = "#FFFFFF"
                        
                        # Handling individual font color for each cell
                        if font.color and font.color.type == 'rgb':
                            font_color = f"#{font.color.rgb[2:]}"
                        else:
                            font_color = "#000000"  # Default font color (black)

                        # Inline CSS for cell: applying font, font size, weight, color, background color, and alignment
                        cell_style = f"""
                            font-family: {font.name or 'Arial'};
                            font-size: {font.sz or 11}px;
                            font-weight: {'bold' if font.bold else 'normal'};
                            font-style: {'italic' if font.italic else 'normal'};
                            text-decoration: {'underline' if font.underline else 'none'};
                            color: {font_color};
                            background-color: {bg_color};
                            text-align: {alignment.horizontal or 'left'};
                            vertical-align: {alignment.vertical or 'top'};
                            border: 1px solid #000;
                        """

                        # Add cell to row HTML
                        row_html += f'<td style="{cell_style}">{cell_value}</td>'
                row_html += '</tr>'
                table_html += row_html

            table_html += '</table>'

            # Add table HTML for the sheet
            sheet_data.append({
                'name': sheet_name,
                'data': table_html
            })

        # Render the page with pagination
        return render_template(
            'display_file.html',
            file_name=file_name,
            sheet_data=sheet_data,
            sheet_count=len(sheet_data),
            pagination_duration=pagination_duration
        )
    except Exception as e:
        return f"Error processing file: {e}", 500

def evaluate_expression(expression):
    """
    Evaluates a mathematical expression string following BODMAS rules.
    Only allows numbers and basic arithmetic operators: +, -, *, /, %, (, ).
    If the expression is not purely mathematical, returns the original string.
    """
    # Remove spaces from the expression
    expression = expression.replace(' ', '')

    # Define a regular expression pattern to match valid mathematical expressions
    valid_pattern = r'^[\d+\-*/%().]+$'

    # Check if the expression contains only valid mathematical characters
    if re.match(valid_pattern, expression):
        try:
            # Evaluate the expression safely
            result = eval(expression)
            return result
        except ZeroDivisionError:
            return "Error: Division by zero."
        except Exception as e:
            return f"Error: {str(e)}"
    else:
        # If not a valid mathematical expression, return the original string
        return expression

def handle_math_operations(value, config_data):
    """
    Handle the math operations inside {{ }} placeholders and replace keys with their values.
    If the key inside {{ }} matches a key in config_data, replace it with the corresponding value.
    """
    math_pattern = r"\{\{([^\}]+)\}\}"  # Pattern to extract {{ ... }}
    matches = re.findall(math_pattern, value)

    for match in matches:
        # Check if the match corresponds to a key in config_data
        if match in config_data.get("result", {}):
            # Replace the placeholder with the value from config_data
            replacement_value = config_data["result"][match]
            value = value.replace(f"{{{{{match}}}}}", str(replacement_value))
        else:
            # If not a direct match, attempt math operations
            components = re.split(r"(\+|\-|\*|\/|\%|\^|\<|\>|\=)", match)
            components = [c.strip() for c in components if c.strip()]

            result = None
            for i in range(0, len(components), 2):
                # Get the left operand (it may be a number or a placeholder)
                left_operand = components[i]
                if left_operand.isdigit():
                    left_value = float(left_operand)
                else:
                    left_value = float(config_data.get("result", {}).get(left_operand, 3))

            # Replace the original {{ }} expression with the calculated result
            value = value.replace(f"{{{{ {match} }}}}", str(result))
            
    # result = evaluate_expression(value)
    return value

@app.route('/update', methods=['GET'])
def update():
    # File name (default example, dynamically change as needed)
    file_name = "UI.xlsx"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)

    if not os.path.exists(file_path):
        return f"File '{file_name}' does not exist on the server.", 404

    # Pagination duration in seconds (you can dynamically pass it via a query parameter)
    pagination_duration = request.args.get('duration', 5, type=int)

    try:
        # Load Excel using openpyxl to handle merged cells
        workbook = load_workbook(file_path, data_only=True)
        sheet_data = []

        # Load config data from the config.json file
        config_data = load_config()
        print("Config : " + str(config_data))
        tries = 3
        while tries !=0:
        # print("Config Data: ",config_data)
            if config_data == {}:
                tries = tries -1
            else:
                tries = 3
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    # Start building HTML for the table with styling
                    table_html = '<table class="table table-bordered" style="border-collapse: collapse;">'

                    # Identify rows and columns that contain data
                    non_empty_rows = []
                    non_empty_columns = set()  # Using a set to avoid duplicate entries

                    # Iterate over rows and check if they contain any data
                    for row in sheet.iter_rows():
                        if any(cell.value not in [None, ''] for cell in row):  # Check if the row contains any data
                            non_empty_rows.append(row[0].row)  # Store row index if it has data

                    # Identify non-empty columns
                    for col in sheet.iter_cols():
                        if any(cell.value not in [None, ''] for cell in col):  # Check if the column contains any data
                            col_index = col[0].column
                            non_empty_columns.add(col_index)  # Add the column index

                    # Generate HTML for each row and column based on non-empty ones
                    row_idx = 1 ;
                    for row in sheet.iter_rows(min_row=min(non_empty_rows), max_row=max(non_empty_rows)):
                        row_html = '<tr>'
                                            # Handle images
                                    
                        for cell in row:
                            col_idx = 1 ;
                            
                            neighbors = {
                                    "left": sheet.cell(row=row_idx, column=col_idx - 1).value if col_idx > 1 else None,
                                    "right": sheet.cell(row=row_idx, column=col_idx + 1).value if col_idx < sheet.max_column else None,
                                    "top": sheet.cell(row=row_idx - 1, column=col_idx).value if row_idx > 1 else None,
                                    "bottom": sheet.cell(row=row_idx + 1, column=col_idx).value if row_idx < sheet.max_row else None
                                }
                            if cell.column in non_empty_columns:  # Only include non-empty columns
                                # Extract cell value
                                value = cell.value

                                # Check if the value is in the format {{key}} and replace it
                                if isinstance(value, str) and value.startswith("{{") and value.endswith("}}"):
                                    temp_val = str(value)
                                    key = value[2:-2].strip()  # Extract the key inside {{}} 
                                    temp_key = str(key)
                                    replacement_value = config_data.get("result", {}).get(temp_key, value)  # Get from config.json or keep original
                                    value = str(replacement_value)
                                    
                                    # Check for mathematical expressions like {{1}}{{+}}{{2}}
                                    # Handle concatenation and operations
                                
                                    value = handle_math_operations(value, config_data)
                                    result = evaluate_expression(str(value))
                                    value = str(result)
                                    
                                # Check for image or video paths in the format {~media_path~}
                                if isinstance(value, str) and value.startswith("{~") and value.endswith("~}"):
                                    media_path = value[2:-2].strip()  # Extract the key inside {~ ~}
                                    full_media_path = os.path.join(app.config['UPLOAD_FOLDER'], media_path)
                                    if os.path.exists(full_media_path):
                                        media_extension = os.path.splitext(media_path)[-1].lower()  # Get file extension
                                        if media_extension in ['.png', '.jpg', '.jpeg', '.gif']:  # Image files
                                            with open(full_media_path, "rb") as media_file:
                                                encoded_media = base64.b64encode(media_file.read()).decode('utf-8')
                                                value = f'<img src="data:image/{media_extension[1:]};base64,{encoded_media}" alt="media" style="width: 100%; height: 100%;">'
                                        elif media_extension in ['.mp4', '.webm', '.ogg']:  # Video files
                                            with open(full_media_path, "rb") as media_file:
                                                encoded_media = base64.b64encode(media_file.read()).decode('utf-8')
                                                value = (
                                                    f'<video autoplay muted loop controls style="width: 100%; height: 100%;">'
                                                    f'<source src="data:video/{media_extension[1:]};base64,{encoded_media}" type="video/{media_extension[1:]}">'
                                                    f'Your browser does not support the video tag.'
                                                    f'</video>'
                                                )
                                        else:
                                            value = "Unsupported media format"
                                    else:
                                        value = "Media not found"


                                # Check for image paths in the format {$image_path$}
                                if isinstance(value, str) and value.startswith("{~") and value.endswith("~}"):
                                    image_path = value[2:-2].strip()  # Extract the key inside {$ $}
                                    full_image_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path)
                                    if os.path.exists(full_image_path):
                                        with open(full_image_path, "rb") as image_file:
                                            encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                                            value = f'<img src="data:image/png;base64,{encoded_image}" alt="image" style="width: 100%; height: 100%">'
                                    else:
                                        value = "Image not found"   
                                        
                                # Replace None with a blank space
                                cell_value = value if value is not None else " "

                                # Extract cell style
                                font = cell.font
                                fill = cell.fill
                                alignment = cell.alignment

                                # Handle all fill color types (rgb, indexed, theme, etc.)
                                if fill.fgColor:
                                    if fill.fgColor.type == "rgb":  # RGB color type
                                        bg_color = f"#{fill.fgColor.rgb[2:]}"  # Extract RGB value (excluding first two "FF")
                                    else:
                                        bg_color = "#FFFFFF"
                                else:
                                    bg_color = "#FFFFFF"
                                
                                # Handling individual font color safely
                                if font.color and font.color.type == 'rgb':
                                    font_color = f"#{font.color.rgb[2:]}"
                                else:
                                    font_color = "#000000"  # Default font color (black)

                                # Handle case when cell is blank (set left/right border as white)
                                if cell_value.strip() == "":
                                        # Determine border logic based on neighbors
                                    if all(val in [None, ''] for val in neighbors.values()):
                                        # No neighbors: Make all borders white
                                        border_style = "border: 1px solid white;"
                                    else:
                                        # Logic for only certain sides of the border
                                        border_styles = []
                                        if not neighbors["left"] and neighbors["right"] and neighbors["top"] and neighbors["bottom"]:
                                            border_styles.append("border-left: 1px solid white;")
                                        if not neighbors["right"] and neighbors["left"] and neighbors["top"] and neighbors["bottom"]:
                                            border_styles.append("border-right: 1px solid white;")
                                        if not neighbors["top"] and neighbors["left"] and neighbors["right"] and neighbors["bottom"]:
                                            border_styles.append("border-top: 1px solid white;")
                                        if not neighbors["bottom"] and neighbors["left"] and neighbors["right"] and neighbors["top"]:
                                            border_styles.append("border-bottom: 1px solid white;")
                                        border_style = " ".join(border_styles)
                                        
                                    cell_style = f"""
                                        font-family: {font.name or 'Arial'};
                                        font-size: {font.sz or 11}px;
                                        font-weight: {'bold' if font.bold else 'normal'};
                                        font-style: {'italic' if font.italic else 'normal'};
                                        text-decoration: {'underline' if font.underline else 'none'};
                                        color: {font_color};
                                        background-color: {bg_color};
                                        text-align: {alignment.horizontal or 'left'};
                                        vertical-align: {alignment.vertical or 'top'};
                                        border: {border_style};
                                    """
                                else:
                                    # Normal style if the cell has content
                                    cell_style = f"""
                                        font-family: {font.name or 'Arial'};
                                        font-size: {font.sz or 11}px;
                                        font-weight: {'bold' if font.bold else 'normal'};
                                        font-style: {'italic' if font.italic else 'normal'};
                                        text-decoration: {'underline' if font.underline else 'none'};
                                        color: {font_color};
                                        background-color: {bg_color};
                                        text-align: {alignment.horizontal or 'left'};
                                        vertical-align: {alignment.vertical or 'top'};
                                        border: 1px solid #000;
                                    """

                                # Add cell to row HTML
                                row_html += f'<td style="{cell_style}">{cell_value}</td>'
                        row_html += '</tr>'
                        col_idx = col_idx + 1
                        table_html += row_html
                    row_idx = row_idx + 1
                    table_html += '</table>'

                    # Add table HTML for the sheet
                    sheet_data.append({
                        'name': sheet_name,
                        'data': table_html
                    })

                return jsonify(sheet_data)
            break
    except Exception as e:
        return f"Error processing file: {e}", 500

@app.route('/', methods=['GET'])
def show():
    # File name (default example, dynamically change as needed)
    file_name = "UI.xlsx"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)

    if not os.path.exists(file_path):
        return f"File '{file_name}' does not exist on the server.", 404

    # Pagination duration in seconds (you can dynamically pass it via a query parameter)
    pagination_duration = request.args.get('duration', 5, type=int)

    try:
        # Load Excel using openpyxl to handle merged cells
        workbook = load_workbook(file_path, data_only=True)
        sheet_data = []

        # Load config data from the config.json file
        config_data = load_config()
        print("Config : " + str(config_data))
        tries = 3
        while tries !=0:
        # print("Config Data: ",config_data)
            if config_data == {}:
                tries = tries -1
            else:
                tries = 3
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    # Start building HTML for the table with styling
                    table_html = '<table class="table table-bordered" style="border-collapse: collapse;">'

                    # Identify rows and columns that contain data
                    non_empty_rows = []
                    non_empty_columns = set()  # Using a set to avoid duplicate entries

                    # Iterate over rows and check if they contain any data
                    for row in sheet.iter_rows():
                        if any(cell.value not in [None, ''] for cell in row):  # Check if the row contains any data
                            non_empty_rows.append(row[0].row)  # Store row index if it has data

                    # Identify non-empty columns
                    for col in sheet.iter_cols():
                        if any(cell.value not in [None, ''] for cell in col):  # Check if the column contains any data
                            col_index = col[0].column
                            non_empty_columns.add(col_index)  # Add the column index

                    # Generate HTML for each row and column based on non-empty ones
                    row_idx = 1 ;
                    for row in sheet.iter_rows(min_row=min(non_empty_rows), max_row=max(non_empty_rows)):
                        row_html = '<tr>'
                                            # Handle images
                                    
                        for cell in row:
                            col_idx = 1 ;
                            
                            neighbors = {
                                    "left": sheet.cell(row=row_idx, column=col_idx - 1).value if col_idx > 1 else None,
                                    "right": sheet.cell(row=row_idx, column=col_idx + 1).value if col_idx < sheet.max_column else None,
                                    "top": sheet.cell(row=row_idx - 1, column=col_idx).value if row_idx > 1 else None,
                                    "bottom": sheet.cell(row=row_idx + 1, column=col_idx).value if row_idx < sheet.max_row else None
                                }
                            if cell.column in non_empty_columns:  # Only include non-empty columns
                                # Extract cell value
                                value = cell.value

                                # Check if the value is in the format {{key}} and replace it
                                if isinstance(value, str) and value.startswith("{{") and value.endswith("}}"):
                                    temp_val = str(value)
                                    key = value[2:-2].strip()  # Extract the key inside {{}} 
                                    temp_key = str(key)
                                    replacement_value = config_data.get("result", {}).get(temp_key, "0")  # Get from config.json or keep original
                                    value = str(replacement_value)
                                    
                                    # Check for mathematical expressions like {{1}}{{+}}{{2}}
                                    # Handle concatenation and operations
                                    # Handle concatenation and operation
                                    value = handle_math_operations(value, config_data)
                                    result = evaluate_expression(str(value))
                                    value = str(result)
                                    
                                # Check for image or video paths in the format {~media_path~}
                                if isinstance(value, str) and value.startswith("{~") and value.endswith("~}"):
                                    media_path = value[2:-2].strip()  # Extract the key inside {~ ~}
                                    full_media_path = os.path.join(app.config['UPLOAD_FOLDER'], media_path)
                                    if os.path.exists(full_media_path):
                                        media_extension = os.path.splitext(media_path)[-1].lower()  # Get file extension
                                        if media_extension in ['.png', '.jpg', '.jpeg', '.gif']:  # Image files
                                            with open(full_media_path, "rb") as media_file:
                                                encoded_media = base64.b64encode(media_file.read()).decode('utf-8')
                                                value = f'<img src="data:image/{media_extension[1:]};base64,{encoded_media}" alt="media" style="width: 100%; height: 100%;">'
                                        elif media_extension in ['.mp4', '.webm', '.ogg']:  # Video files
                                            with open(full_media_path, "rb") as media_file:
                                                encoded_media = base64.b64encode(media_file.read()).decode('utf-8')
                                                value = (
                                                    f'<video autoplay muted loop controls style="width: 100%; height: 100%;">'
                                                    f'<source src="data:video/{media_extension[1:]};base64,{encoded_media}" type="video/{media_extension[1:]}">'
                                                    f'Your browser does not support the video tag.'
                                                    f'</video>'
                                                )
                                        else:
                                            value = "Unsupported media format"
                                    else:
                                        value = "Media not found"


                                # Check for image paths in the format {$image_path$}
                                if isinstance(value, str) and value.startswith("{~") and value.endswith("~}"):
                                    image_path = value[2:-2].strip()  # Extract the key inside {$ $}
                                    full_image_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path)
                                    if os.path.exists(full_image_path):
                                        with open(full_image_path, "rb") as image_file:
                                            encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                                            value = f'<img src="data:image/png;base64,{encoded_image}" alt="image" style="width: 100%; height: 100%">'
                                    else:
                                        value = "Image not found"   
                                
                                # # Handle scrolling text logic using CSS animations
                                # if isinstance(value, str) and value.startswith("{@") and value.endswith("@}"):
                                #     print("Scrolling Found=============================================")
                                #     try:
                                #         message = "Hello"
                                #         speed_type = 2  # Change this value for LTR, RTL, or static behavior
                                #         speed_value = 1  # Speed of scrolling

                                #         # Generate appropriate CSS animation styles
                                #         if speed_type == 1:
                                #             # Static message
                                #             value = f'<div style="overflow: hidden;">{message}</div>'
                                #         elif speed_type == 2:
                                #             # Left-to-right scroll using CSS animation
                                #             value = f'''
                                #                 <div class="scroll-container-ltr" style="overflow: hidden;">
                                #                     <div class="scroll-text-ltr">{message}</div>
                                #                 </div>
                                #                 <style>
                                #                 @keyframes scrollLeftToRight {{
                                #                     0% {{ transform: translateX(100%); }}
                                #                     100% {{ transform: translateX(-100%); }}
                                #                 }}
                                #                 .scroll-container-ltr {{
                                #                     white-space: nowrap;
                                #                     overflow: hidden;
                                #                 }}
                                #                 .scroll-text-ltr {{
                                #                     display: inline-block;
                                #                     animation: scrollLeftToRight {speed_value}s linear infinite;
                                #                 }}
                                #                 </style>
                                #             '''
                                #         elif speed_type == 3:
                                #             # Right-to-left scroll using CSS animation
                                #             value = f'''
                                #                 <div class="scroll-container-rtl" style="overflow: hidden;">
                                #                     <div class="scroll-text-rtl">{message}</div>
                                #                 </div>
                                #                 <style>
                                #                 @keyframes scrollRightToLeft {{
                                #                     0% {{ transform: translateX(-100%); }}
                                #                     100% {{ transform: translateX(100%); }}
                                #                 }}
                                #                 .scroll-container-rtl {{
                                #                     white-space: nowrap;
                                #                     overflow: hidden;
                                #                 }}
                                #                 .scroll-text-rtl {{
                                #                     display: inline-block;
                                #                     animation: scrollRightToLeft {speed_value}s linear infinite;
                                #                 }}
                                #                 </style>
                                #             '''
                                #     except ValueError:
                                #         value = "Invalid message format"
                                # Replace None with a blank space
                                cell_value = value if value is not None else " "

                                # Extract cell style
                                font = cell.font
                                fill = cell.fill
                                alignment = cell.alignment

                                # Handle all fill color types (rgb, indexed, theme, etc.)
                                if fill.fgColor:
                                    if fill.fgColor.type == "rgb":  # RGB color type
                                        bg_color = f"#{fill.fgColor.rgb[2:]}"  # Extract RGB value (excluding first two "FF")
                                    else:
                                        bg_color = "#FFFFFF"
                                else:
                                    bg_color = "#FFFFFF"
                                
                                # Handling individual font color safely
                                if font.color and font.color.type == 'rgb':
                                    font_color = f"#{font.color.rgb[2:]}"
                                else:
                                    font_color = "#000000"  # Default font color (black)

                                # Handle case when cell is blank (set left/right border as white)
                                if cell_value.strip() == "":
                                        # Determine border logic based on neighbors
                                    if all(val in [None, ''] for val in neighbors.values()):
                                        # No neighbors: Make all borders white
                                        border_style = "border: 1px solid white;"
                                    else:
                                        # Logic for only certain sides of the border
                                        border_styles = []
                                        if not neighbors["left"] and neighbors["right"] and neighbors["top"] and neighbors["bottom"]:
                                            border_styles.append("border-left: 1px solid white;")
                                        if not neighbors["right"] and neighbors["left"] and neighbors["top"] and neighbors["bottom"]:
                                            border_styles.append("border-right: 1px solid white;")
                                        if not neighbors["top"] and neighbors["left"] and neighbors["right"] and neighbors["bottom"]:
                                            border_styles.append("border-top: 1px solid white;")
                                        if not neighbors["bottom"] and neighbors["left"] and neighbors["right"] and neighbors["top"]:
                                            border_styles.append("border-bottom: 1px solid white;")
                                        border_style = " ".join(border_styles)
                                        
                                    cell_style = f"""
                                        font-family: {font.name or 'Arial'};
                                        font-size: {font.sz or 11}px;
                                        font-weight: {'bold' if font.bold else 'normal'};
                                        font-style: {'italic' if font.italic else 'normal'};
                                        text-decoration: {'underline' if font.underline else 'none'};
                                        color: {font_color};
                                        background-color: {bg_color};
                                        text-align: {alignment.horizontal or 'left'};
                                        vertical-align: {alignment.vertical or 'top'};
                                        border: {border_style};
                                    """
                                else:
                                    # Normal style if the cell has content
                                    cell_style = f"""
                                        font-family: {font.name or 'Arial'};
                                        font-size: {font.sz or 11}px;
                                        font-weight: {'bold' if font.bold else 'normal'};
                                        font-style: {'italic' if font.italic else 'normal'};
                                        text-decoration: {'underline' if font.underline else 'none'};
                                        color: {font_color};
                                        background-color: {bg_color};
                                        text-align: {alignment.horizontal or 'left'};
                                        vertical-align: {alignment.vertical or 'top'};
                                        border: 1px solid #000;
                                    """

                                # Add cell to row HTML
                                row_html += f'<td style="{cell_style}">{cell_value}</td>'
                        row_html += '</tr>'
                        col_idx = col_idx + 1
                        table_html += row_html
                    row_idx = row_idx + 1
                    table_html += '</table>'

                    # Add table HTML for the sheet
                    sheet_data.append({
                        'name': sheet_name,
                        'data': table_html
                    })
                    
                print("Pages dur=========================================================================")
                # Array to store extracted numbers
                extracted_numbers = []

                # Process each page
                global pages
                trimmed_values = [page[2:-2] for page in pages]
                print("Pages:",trimmed_values )
                # Render the page with pagination
                return render_template(
                    'show.html',
                    file_name=file_name,
                    sheet_data=sheet_data,
                    sheet_count=len(sheet_data),
                    pagination_duration=pagination_duration,
                    pages=trimmed_values   # Send the pages array from config to the template
                )
                break
    except Exception as e:
        return f"Error processing file: {e}", 500

# @app.route('/', methods=['GET'])
# def show():
#     # File name (default example, dynamically change as needed)
#     file_name = "Book2.xlsx"
#     file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)

#     if not os.path.exists(file_path):
#         return f"File '{file_name}' does not exist on the server.", 404

#     # Pagination duration in seconds (you can dynamically pass it via a query parameter)
#     pagination_duration = request.args.get('duration', 5, type=int)

#     try:
#         # Load Excel using openpyxl to handle merged cells
#         workbook = load_workbook(file_path, data_only=True)
#         sheet_data = []

#         # Load config data from the config.json file
#         config_data = load_config()

#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]

#             # Start building HTML for the table with styling
#             table_html = '<table class="table table-bordered" style="border-collapse: collapse;">'

#             # Identify rows and columns that contain data
#             non_empty_rows = []
#             non_empty_columns = set()  # Using a set to avoid duplicate entries

#             # Iterate over rows and check if they contain any data
#             for row in sheet.iter_rows():
#                 if any(cell.value not in [None, ''] for cell in row):  # Check if the row contains any data
#                     non_empty_rows.append(row[0].row)  # Store row index if it has data

#             # Identify non-empty columns
#             for col in sheet.iter_cols():
#                 if any(cell.value not in [None, ''] for cell in col):  # Check if the column contains any data
#                     col_index = col[0].column
#                     non_empty_columns.add(col_index)  # Add the column index

#             # Generate HTML for each row and column based on non-empty ones
#             row_idx = 1 ;
#             for row in sheet.iter_rows(min_row=min(non_empty_rows), max_row=max(non_empty_rows)):
#                 row_html = '<tr>'
#                                     # Handle images
                            
#                 for cell in row:
#                     col_idx = 1 ;
                    
#                     neighbors = {
#                             "left": sheet.cell(row=row_idx, column=col_idx - 1).value if col_idx > 1 else None,
#                             "right": sheet.cell(row=row_idx, column=col_idx + 1).value if col_idx < sheet.max_column else None,
#                             "top": sheet.cell(row=row_idx - 1, column=col_idx).value if row_idx > 1 else None,
#                             "bottom": sheet.cell(row=row_idx + 1, column=col_idx).value if row_idx < sheet.max_row else None
#                         }
#                     if cell.column in non_empty_columns:  # Only include non-empty columns
#                         # Extract cell value
#                         value = cell.value

#                         # Check if the value is in the format {{key}} and replace it
#                         if isinstance(value, str) and value.startswith("{{") and value.endswith("}}"):
#                             key = value[2:-2].strip()  # Extract the key inside {{}} 
#                             replacement_value = config_data.get("result", {}).get(key, value)  # Get from config.json or keep original
#                             value = replacement_value
                            
#                         # Check for image or video paths in the format {~media_path~}
#                         if isinstance(value, str) and value.startswith("{~") and value.endswith("~}"):
#                             media_path = value[2:-2].strip()  # Extract the key inside {~ ~}
#                             full_media_path = os.path.join(app.config['UPLOAD_FOLDER'], media_path)
#                             if os.path.exists(full_media_path):
#                                 media_extension = os.path.splitext(media_path)[-1].lower()  # Get file extension
#                                 if media_extension in ['.png', '.jpg', '.jpeg', '.gif']:  # Image files
#                                     with open(full_media_path, "rb") as media_file:
#                                         encoded_media = base64.b64encode(media_file.read()).decode('utf-8')
#                                         value = f'<img src="data:image/{media_extension[1:]};base64,{encoded_media}" alt="media" style="width: 100%; height: 100%;">'
#                                 elif media_extension in ['.mp4', '.webm', '.ogg']:  # Video files
#                                     with open(full_media_path, "rb") as media_file:
#                                         encoded_media = base64.b64encode(media_file.read()).decode('utf-8')
#                                         value = (
#                                             f'<video autoplay muted loop controls style="width: 100%; height: 100%;">'
#                                             f'<source src="data:video/{media_extension[1:]};base64,{encoded_media}" type="video/{media_extension[1:]}">'
#                                             f'Your browser does not support the video tag.'
#                                             f'</video>'
#                                         )
#                                 else:
#                                     value = "Unsupported media format"
#                             else:
#                                 value = "Media not found"


#                         # Check for image paths in the format {$image_path$}
#                         if isinstance(value, str) and value.startswith("{~") and value.endswith("~}"):
#                             image_path = value[2:-2].strip()  # Extract the key inside {$ $}
#                             full_image_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path)
#                             if os.path.exists(full_image_path):
#                                 with open(full_image_path, "rb") as image_file:
#                                     encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
#                                     value = f'<img src="data:image/png;base64,{encoded_image}" alt="image" style="width: 100%; height: 100%">'
#                             else:
#                                 value = "Image not found"                    

#                         # Replace None with a blank space
#                         cell_value = value if value is not None else " "

#                         # Extract cell style
#                         font = cell.font
#                         fill = cell.fill
#                         alignment = cell.alignment

#                         # Handle all fill color types (rgb, indexed, theme, etc.)
#                         if fill.fgColor:
#                             if fill.fgColor.type == "rgb":  # RGB color type
#                                 bg_color = f"#{fill.fgColor.rgb[2:]}"  # Extract RGB value (excluding first two "FF")
#                             else:
#                                 bg_color = "#FFFFFF"
#                         else:
#                             bg_color = "#FFFFFF"
                        
#                         # Handling individual font color safely
#                         if font.color and font.color.type == 'rgb':
#                             font_color = f"#{font.color.rgb[2:]}"
#                         else:
#                             font_color = "#000000"  # Default font color (black)

#                         # Handle case when cell is blank (set left/right border as white)
#                         if cell_value.strip() == "":
#                                 # Determine border logic based on neighbors
#                             if all(val in [None, ''] for val in neighbors.values()):
#                                 # No neighbors: Make all borders white
#                                 border_style = "border: 1px solid white;"
#                             else:
#                                 # Logic for only certain sides of the border
#                                 border_styles = []
#                                 if not neighbors["left"] and neighbors["right"] and neighbors["top"] and neighbors["bottom"]:
#                                     border_styles.append("border-left: 1px solid white;")
#                                 if not neighbors["right"] and neighbors["left"] and neighbors["top"] and neighbors["bottom"]:
#                                     border_styles.append("border-right: 1px solid white;")
#                                 if not neighbors["top"] and neighbors["left"] and neighbors["right"] and neighbors["bottom"]:
#                                     border_styles.append("border-top: 1px solid white;")
#                                 if not neighbors["bottom"] and neighbors["left"] and neighbors["right"] and neighbors["top"]:
#                                     border_styles.append("border-bottom: 1px solid white;")
#                                 border_style = " ".join(border_styles)
                                
#                             cell_style = f"""
#                                 font-family: {font.name or 'Arial'};
#                                 font-size: {font.sz or 11}px;
#                                 font-weight: {'bold' if font.bold else 'normal'};
#                                 font-style: {'italic' if font.italic else 'normal'};
#                                 text-decoration: {'underline' if font.underline else 'none'};
#                                 color: {font_color};
#                                 background-color: {bg_color};
#                                 text-align: {alignment.horizontal or 'left'};
#                                 vertical-align: {alignment.vertical or 'top'};
#                                 border: {border_style};
#                             """
#                         else:
#                             # Normal style if the cell has content
#                             cell_style = f"""
#                                 font-family: {font.name or 'Arial'};
#                                 font-size: {font.sz or 11}px;
#                                 font-weight: {'bold' if font.bold else 'normal'};
#                                 font-style: {'italic' if font.italic else 'normal'};
#                                 text-decoration: {'underline' if font.underline else 'none'};
#                                 color: {font_color};
#                                 background-color: {bg_color};
#                                 text-align: {alignment.horizontal or 'left'};
#                                 vertical-align: {alignment.vertical or 'top'};
#                                 border: 1px solid #000;
#                             """

#                         # Add cell to row HTML
#                         row_html += f'<td style="{cell_style}">{cell_value}</td>'
#                 row_html += '</tr>'
#                 col_idx = col_idx + 1
#                 table_html += row_html
#             row_idx = row_idx + 1
#             table_html += '</table>'

#             # Add table HTML for the sheet
#             sheet_data.append({
#                 'name': sheet_name,
#                 'data': table_html
#             })

#         # Render the page with pagination
#         return render_template(
#             'show.html',
#             file_name=file_name,
#             sheet_data=sheet_data,
#             sheet_count=len(sheet_data),
#             pagination_duration=pagination_duration
#         )
#     except Exception as e:
#         return f"Error processing file: {e}", 500

# ---  last save working
# @app.route('/', methods=['GET'])
# def show():
#     # File name (default example, dynamically change as needed)
#     file_name = "Book2.xlsx"
#     file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)

#     if not os.path.exists(file_path):
#         return f"File '{file_name}' does not exist on the server.", 404

#     # Pagination duration in seconds (you can dynamically pass it via a query parameter)
#     pagination_duration = request.args.get('duration', 5, type=int)

#     try:
#         # Load Excel using openpyxl to handle merged cells
#         workbook = load_workbook(file_path, data_only=True)
#         sheet_data = []

#         # Load config data from the config.json file
#         config_data = load_config()

#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]

#             # Start building HTML for the table with styling
#             table_html = '<table class="table table-bordered" style="border-collapse: collapse;">'

#             # Identify rows and columns that contain data
#             non_empty_rows = []
#             non_empty_columns = set()  # Using a set to avoid duplicate entries

#             # Iterate over rows and check if they contain any data
#             for row in sheet.iter_rows():
#                 if any(cell.value not in [None, ''] for cell in row):  # Check if the row contains any data
#                     non_empty_rows.append(row[0].row)  # Store row index if it has data

#             # Identify non-empty columns
#             for col in sheet.iter_cols():
#                 if any(cell.value not in [None, ''] for cell in col):  # Check if the column contains any data
#                     col_index = col[0].column
#                     non_empty_columns.add(col_index)  # Add the column index

#             # Generate HTML for each row and column based on non-empty ones
#             row_idx = 1 ;
#             for row in sheet.iter_rows(min_row=min(non_empty_rows), max_row=max(non_empty_rows)):
#                 row_html = '<tr>'
                
#                 for cell in row:
#                     col_idx = 1 ;
#                     neighbors = {
#                             "left": sheet.cell(row=row_idx, column=col_idx - 1).value if col_idx > 1 else None,
#                             "right": sheet.cell(row=row_idx, column=col_idx + 1).value if col_idx < sheet.max_column else None,
#                             "top": sheet.cell(row=row_idx - 1, column=col_idx).value if row_idx > 1 else None,
#                             "bottom": sheet.cell(row=row_idx + 1, column=col_idx).value if row_idx < sheet.max_row else None
#                         }
#                     if cell.column in non_empty_columns:  # Only include non-empty columns
#                         # Extract cell value
#                         value = cell.value

#                         # Check if the value is in the format {{key}} and replace it
#                         if isinstance(value, str) and value.startswith("{{") and value.endswith("}}"):
#                             key = value[2:-2].strip()  # Extract the key inside {{}} 
#                             replacement_value = config_data.get("result", {}).get(key, value)  # Get from config.json or keep original
#                             value = replacement_value

#                         # Replace None with a blank space
#                         cell_value = value if value is not None else " "

#                         # Extract cell style
#                         font = cell.font
#                         fill = cell.fill
#                         alignment = cell.alignment

#                         # Handle all fill color types (rgb, indexed, theme, etc.)
#                         if fill.fgColor:
#                             if fill.fgColor.type == "rgb":  # RGB color type
#                                 bg_color = f"#{fill.fgColor.rgb[2:]}"  # Extract RGB value (excluding first two "FF")
#                             else:
#                                 bg_color = "#FFFFFF"
#                         else:
#                             bg_color = "#FFFFFF"
                        
#                         # Handling individual font color safely
#                         if font.color and font.color.type == 'rgb':
#                             font_color = f"#{font.color.rgb[2:]}"
#                         else:
#                             font_color = "#000000"  # Default font color (black)

#                         # Handle case when cell is blank (set left/right border as white)
#                         if cell_value.strip() == "":
#                                 # Determine border logic based on neighbors
#                             if all(val in [None, ''] for val in neighbors.values()):
#                                 # No neighbors: Make all borders white
#                                 border_style = "border: 1px solid white;"
#                             else:
#                                 # Logic for only certain sides of the border
#                                 border_styles = []
#                                 if not neighbors["left"] and neighbors["right"] and neighbors["top"] and neighbors["bottom"]:
#                                     border_styles.append("border-left: 1px solid white;")
#                                 if not neighbors["right"] and neighbors["left"] and neighbors["top"] and neighbors["bottom"]:
#                                     border_styles.append("border-right: 1px solid white;")
#                                 if not neighbors["top"] and neighbors["left"] and neighbors["right"] and neighbors["bottom"]:
#                                     border_styles.append("border-top: 1px solid white;")
#                                 if not neighbors["bottom"] and neighbors["left"] and neighbors["right"] and neighbors["top"]:
#                                     border_styles.append("border-bottom: 1px solid white;")
#                                 border_style = " ".join(border_styles)
                                
#                             cell_style = f"""
#                                 font-family: {font.name or 'Arial'};
#                                 font-size: {font.sz or 11}px;
#                                 font-weight: {'bold' if font.bold else 'normal'};
#                                 font-style: {'italic' if font.italic else 'normal'};
#                                 text-decoration: {'underline' if font.underline else 'none'};
#                                 color: {font_color};
#                                 background-color: {bg_color};
#                                 text-align: {alignment.horizontal or 'left'};
#                                 vertical-align: {alignment.vertical or 'top'};
#                                 border: {border_style};
#                             """
#                         else:
#                             # Normal style if the cell has content
#                             cell_style = f"""
#                                 font-family: {font.name or 'Arial'};
#                                 font-size: {font.sz or 11}px;
#                                 font-weight: {'bold' if font.bold else 'normal'};
#                                 font-style: {'italic' if font.italic else 'normal'};
#                                 text-decoration: {'underline' if font.underline else 'none'};
#                                 color: {font_color};
#                                 background-color: {bg_color};
#                                 text-align: {alignment.horizontal or 'left'};
#                                 vertical-align: {alignment.vertical or 'top'};
#                                 border: 1px solid #000;
#                             """

#                         # Add cell to row HTML
#                         row_html += f'<td style="{cell_style}">{cell_value}</td>'
#                 row_html += '</tr>'
#                 col_idx = col_idx + 1
#                 table_html += row_html
#             row_idx = row_idx + 1
#             table_html += '</table>'

#             # Add table HTML for the sheet
#             sheet_data.append({
#                 'name': sheet_name,
#                 'data': table_html
#             })

#         # Render the page with pagination
#         return render_template(
#             'show.html',
#             file_name=file_name,
#             sheet_data=sheet_data,
#             sheet_count=len(sheet_data),
#             pagination_duration=pagination_duration
#         )
#     except Exception as e:
#         return f"Error processing file: {e}", 500

# @app.route('/check-file', methods=['GET'])
# def check_file():
#     # File name (default example, dynamically change as needed)
#     file_name = "Book2.xlsx"
#     file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)

#     if not os.path.exists(file_path):
#         return f"File '{file_name}' does not exist on the server.", 404

#     # Pagination duration in seconds (you can dynamically pass it via a query parameter)
#     pagination_duration = request.args.get('duration', 5, type=int)

#     try:
#         # Load Excel using openpyxl to handle merged cells
#         workbook = load_workbook(file_path, data_only=True)
#         sheet_data = []

#         # Load config data from the config.json file
#         config_data = load_config()

#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]

#             # Start building HTML for the table with styling
#             table_html = '<table class="table table-bordered" style="border-collapse: collapse;">'

#             # Identify rows and columns that contain data
#             non_empty_rows = []
#             non_empty_columns = set()  # Using a set to avoid duplicate entries

#             # Store images found in the sheet
#             images = []

#             # Iterate over rows and check if they contain any data
#             for row in sheet.iter_rows():
#                 if any(cell.value not in [None, ''] for cell in row):  # Check if the row contains any data
#                     non_empty_rows.append(row[0].row)  # Store row index if it has data

#             # Identify non-empty columns
#             for col in sheet.iter_cols():
#                 if any(cell.value not in [None, ''] for cell in col):  # Check if the column contains any data
#                     col_index = col[0].column
#                     non_empty_columns.add(col_index)  # Add the column index

#             # Process images in the sheet
#             for image in sheet._images:
#                 if isinstance(image, Image):
#                     # Convert image to base64 to embed in HTML
#                     print("################### === #####################")
#                     img_stream = BytesIO()
#                     image.image.save(img_stream, format='PNG')  # Save image to a stream
#                     img_base64 = base64.b64encode(img_stream.getvalue()).decode('utf-8')
#                     img_tag = f'<img src="data:image/png;base64,{img_base64}" alt="Image" style="max-width: 100px; max-height: 100px;"/>'
#                     images.append(img_tag)

#             # Generate HTML for each row and column based on non-empty ones
#             for row in sheet.iter_rows(min_row=min(non_empty_rows), max_row=max(non_empty_rows)):
#                 row_html = '<tr>'
#                 for cell in row:
#                     if cell.column in non_empty_columns:  # Only include non-empty columns
#                         # Extract cell value
#                         value = cell.value

#                         # Check if the value is in the format {{key}} and replace it
#                         if isinstance(value, str) and value.startswith("{{") and value.endswith("}}"):
#                             key = value[2:-2].strip()  # Extract the key inside {{}} 
#                             replacement_value = config_data.get("result", {}).get(key, value)  # Get from config.json or keep original
#                             value = replacement_value

#                         # Replace None with a blank space
#                         cell_value = value if value is not None else " "

#                         # Extract cell style
#                         font = cell.font
#                         fill = cell.fill
#                         alignment = cell.alignment

#                         # Handle all fill color types (rgb, indexed, theme, etc.)
#                         if fill.fgColor:
#                             if fill.fgColor.type == "rgb":  # RGB color type
#                                 bg_color = f"#{fill.fgColor.rgb[2:]}"  # Extract RGB value (excluding first two "FF")
#                             elif fill.fgColor.type == "indexed":  # Indexed color type
#                                 bg_color = fill.fgColor.indexed
#                             elif fill.fgColor.type == "theme":  # Theme color type
#                                 theme_color = fill.fgColor.theme
#                                 bg_color = f"theme-{theme_color}"  # Applying default "theme" label
#                             else:
#                                 bg_color = "#FFFFFF"
#                         else:
#                             bg_color = "#FFFFFF"
                        
#                         # Handling individual font color for each cell
#                         if font.color and font.color.type == 'rgb':
#                             font_color = f"#{font.color.rgb[2:]}"
#                         else:
#                             font_color = "#000000"  # Default font color (black)

#                         # Inline CSS for cell: applying font, font size, weight, color, background color, and alignment
#                         cell_style = f"""
#                             font-family: {font.name or 'Arial'};
#                             font-size: {font.sz or 11}px;
#                             font-weight: {'bold' if font.bold else 'normal'};
#                             font-style: {'italic' if font.italic else 'normal'};
#                             text-decoration: {'underline' if font.underline else 'none'};
#                             color: {font_color};
#                             background-color: {bg_color};
#                             text-align: {alignment.horizontal or 'left'};
#                             vertical-align: {alignment.vertical or 'top'};
#                             border: 1px solid #000;
#                         """

#                         # Add cell to row HTML
#                         row_html += f'<td style="{cell_style}">{cell_value}</td>'
#                 row_html += '</tr>'
#                 table_html += row_html

#             table_html += '</table>'

#             # Add table HTML for the sheet and images to the data
#             sheet_data.append({
#                 'name': sheet_name,
#                 'data': table_html,
#                 'images': images  # Store images here
#             })

#         # Render the page with pagination
#         return render_template(
#             'display_file.html',
#             file_name=file_name,
#             sheet_data=sheet_data,
#             sheet_count=len(sheet_data),
#             pagination_duration=pagination_duration
#         )
#     except Exception as e:
#         return f"Error processing file: {e}", 500

# Global dictionary to store counter access logs
counter_log = {}
counter_status = {}  # Stores counters with their last updated timestamp

@app.route("/counter", methods=["GET"])
def update_counter_get():
    """Update the counter value in the JSON file using a GET request and log access time."""
    global counter_log, counter_status  # Ensure we modify global objects

    counter_id = request.args.get("counter_id", type=int)
    new_value = request.args.get("value", type=int)

    if counter_id is None or new_value is None:
        return jsonify({"error": "Missing counter_id or value"}), 400

    try:
        # Read the current JSON data
        with open(jsonFilePath, "r", encoding="utf-8") as jsonf:
            data = json.load(jsonf)

        # Update the value if the counter ID exists
        if "result" in data and str(counter_id) in data["result"]:
            data["result"][str(counter_id)] = new_value
            
            # Get the current timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Update both logs
            counter_log[str(counter_id)] = timestamp
            counter_status[str(counter_id)] = timestamp  # Update counter_status
            
            # Write the updated data back to the JSON file
            with open(jsonFilePath, "w", encoding="utf-8") as jsonf:
                json.dump(data, jsonf, indent=4)

            return jsonify({
                "message": f"Counter {counter_id} updated to {new_value}",
                "log": counter_log,
                "status": counter_status
            }), 200
        else:
            return jsonify({"error": f"Counter {counter_id} not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route("/counter-status", methods=["GET"])
def get_counter_status():
    """Check the status of counters based on the last update time."""
    global counter_status  # Ensure we modify the global object

    status_report = {}
    current_time = datetime.now()

    for counter_id, timestamp_str in counter_status.items():
        try:
            # Convert timestamp string to datetime object
            last_update_time = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
            time_diff = current_time - last_update_time

            # Set status based on time difference
            status_report[counter_id] = "Active" if time_diff <= timedelta(minutes=2) else "Inactive"

        except Exception as e:
            print(f"Error processing counter {counter_id}: {e}")
            status_report[counter_id] = "Inactive"

    return jsonify({"counter_status": status_report}), 200


def main():
    log_filename = setup_logging()
    global STOPThread
    global my_thread
    
    try:
        # Log script start
        logging.info("Script started.")
        print(Fore.YELLOW + "Script started.")

        config_path = "CONFIG/CONFIG.txt"
        lines = read_config(config_path)
        config_data = parse_config(lines)

        # Log settings
        logging.info(f"Parsed Configuration: {config_data}")
        print(Fore.YELLOW + f"Version: {config_data['version']}")
        print(Fore.YELLOW + f"Company Name: {config_data['company_name']}")

        print(Fore.YELLOW + "\n--- Counter Settings ---")
        for idx, (cindex, cid) in enumerate(config_data['counters'], 1):
            print(Fore.YELLOW + f"Counter {idx}: CIndex={cindex}, CID={cid}")
            logging.info(f"Counter {idx}: CIndex={cindex}, CID={cid}")

        print(Fore.YELLOW + "\n--- CSV Settings ---")
        for idx, path in enumerate(config_data['csv_paths'], 1):
            print(Fore.YELLOW + f"CSV Path {idx}: {path}")
            logging.info(f"CSV Path {idx}: {path}")

        print(Fore.YELLOW + "\n--- Display Settings ---")
        for idx, line in enumerate(config_data['lines_data'], 1):
            print(Fore.YELLOW + f"Line {idx}: {line}")
            logging.info(f"Line {idx}: {line}")

        print(Fore.YELLOW + f"Log Interval: {config_data['log_interval']}")
        logging.info(f"Log Interval: {config_data['log_interval']}")

        print(Fore.YELLOW + "\n--- Page Control ---")
        for idx, page in enumerate(config_data['pages'], 1):
            print(Fore.YELLOW + f"Page {idx}: {page}")
            logging.info(f"Page {idx}: {page}")
        
        my_thread = threading.Timer(1, Independent, [config_data]).start()
        # Run Flask App
        app.run(host="0.0.0.0", port=5000)
        # Log script termination
        # logging.info("Script terminated successfully.")
        # print(Fore.YELLOW + "Script terminated.")

    except Exception as e:
        logging.error(f"Error: {e}")
        print(Fore.RED + f"Error: {e}")
        logging.info("Script terminated successfully.")
        print(Fore.YELLOW + "Script terminated.")

if __name__ == "__main__":
    main()

# Developed By Girish Pawar & Vishal Padyal